"""
Gerenciador de sincronização de dados.

Responsável por:
- Baixar planilha do Google Sheets
- Verificar alterações via hash MD5
- Processar dados
- Salvar no banco de dados
"""

import hashlib
import re
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import settings
from core.database import Database


class SyncManager:
    """Gerenciador de sincronização."""
    
    def __init__(self):
        self.db = Database()
        self.arquivo_excel: Optional[Path] = None
        self.dados_processados: List[Dict] = []
        self.abas_processadas: List[Dict] = []
    
    # ==================== DOWNLOAD ====================
    
    def _extrair_sheet_id(self, url: str) -> Optional[str]:
        """Extrai o ID da planilha da URL."""
        patterns = [
            r'/spreadsheets/d/([a-zA-Z0-9-_]+)',
            r'id=([a-zA-Z0-9-_]+)',
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None
    
    def baixar_planilha(self, forcar: bool = False) -> Optional[Path]:
        """
        Baixa a planilha do Google Sheets.
        
        Args:
            forcar: Se True, baixa mesmo que já exista arquivo recente
            
        Returns:
            Caminho do arquivo baixado ou None
        """
        print("📥 Verificando planilha...")
        
        # Verifica arquivo existente
        arquivos = list(settings.DOWNLOAD_DIR.glob("planilha_*.xlsx"))
        if arquivos and not forcar:
            arquivo_recente = max(arquivos, key=lambda x: x.stat().st_mtime)
            idade_min = (datetime.now().timestamp() - arquivo_recente.stat().st_mtime) / 60
            
            if idade_min < settings.CACHE_MINUTES:
                print(f"   ✅ Usando cache: {arquivo_recente.name} ({int(idade_min)} min)")
                self.arquivo_excel = arquivo_recente
                return arquivo_recente
        
        # Baixa nova planilha
        print("   ⬇️  Baixando do Google Sheets...")
        
        try:
            sheet_id = self._extrair_sheet_id(settings.GOOGLE_SHEETS_URL)
            if not sheet_id:
                raise ValueError("URL inválida do Google Sheets")
            
            excel_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"planilha_{timestamp}.xlsx"
            caminho = settings.DOWNLOAD_DIR / nome_arquivo
            
            urllib.request.urlretrieve(excel_url, caminho)
            
            print(f"   ✅ Baixado: {nome_arquivo}")
            self.arquivo_excel = caminho
            
            # Limpa antigos
            self._limpar_arquivos_antigos()
            
            return caminho
            
        except Exception as e:
            print(f"   ❌ Erro: {e}")
            return None
    
    def _limpar_arquivos_antigos(self, manter: int = 3):
        """Remove arquivos antigos."""
        arquivos = sorted(
            settings.DOWNLOAD_DIR.glob("planilha_*.xlsx"),
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )
        for arquivo in arquivos[manter:]:
            try:
                arquivo.unlink()
                print(f"   🗑️  Removido: {arquivo.name}")
            except:
                pass
    
    # ==================== HASH / VERIFICAÇÃO ====================
    
    def calcular_hash(self, arquivo: Path = None) -> str:
        """Calcula hash MD5 do arquivo."""
        arquivo = arquivo or self.arquivo_excel
        if not arquivo or not arquivo.exists():
            return ""
        
        with open(arquivo, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()
    
    def arquivo_mudou(self, novo_hash: str) -> bool:
        """Verifica se o arquivo mudou desde última sync."""
        if not settings.HASH_FILE.exists():
            return True
        
        hash_anterior = settings.HASH_FILE.read_text().strip()
        return hash_anterior != novo_hash
    
    def salvar_hash(self, hash_value: str):
        """Salva hash para próxima comparação."""
        settings.HASH_FILE.write_text(hash_value)
    
    # ==================== PROCESSAMENTO ====================
    
    def _corrigir_data(self, dt: datetime, data_saida: datetime = None) -> datetime:
        """Corrige inversão dia/mês em datas."""
        if not isinstance(dt, datetime):
            return dt
        
        dia, mes, ano = dt.day, dt.month, dt.year
        
        if dia > 12:
            return dt
        
        try:
            dt_invertido = datetime(ano, dia, mes)
        except ValueError:
            return dt
        
        # Se temos data de saída para comparar
        if data_saida and isinstance(data_saida, datetime):
            if dt < data_saida and dt_invertido > data_saida:
                return dt_invertido
            if dt > data_saida and dt_invertido > data_saida:
                diff_orig = (dt - data_saida).days
                diff_inv = (dt_invertido - data_saida).days
                if diff_orig > 60 and diff_inv < 45:
                    return dt_invertido
        
        # Heurística: dia=01 e mês>1
        if dia == 1 and 1 < mes <= 12:
            return dt_invertido
        
        return dt
    
    def _validar_data_contexto(self, dt: datetime, mes_aba: int, ano_aba: int) -> datetime:
        """
        Valida e corrige data baseado no contexto da aba.
        
        Se a data não faz sentido para a aba (ex: data em janeiro numa aba de dezembro),
        tenta inverter dia/mês para ver se faz mais sentido.
        
        Args:
            dt: Data a validar
            mes_aba: Mês da aba (1-12)
            ano_aba: Ano da aba
            
        Returns:
            Data corrigida ou original
        """
        if not isinstance(dt, datetime) or not mes_aba:
            return dt
        
        dia, mes = dt.day, dt.month
        
        # Se dia > 12, não tem como inverter
        if dia > 12:
            return dt
        
        # Se já está no mês correto da aba, não precisa corrigir
        if mes == mes_aba:
            return dt
        
        # Tenta inverter dia/mês
        try:
            dt_invertido = datetime(dt.year, dia, mes)
        except ValueError:
            return dt
        
        # Se a versão invertida cair no mês da aba, usa ela
        if dt_invertido.month == mes_aba:
            return dt_invertido
        
        # Se nenhuma das duas é o mês da aba, verifica qual está mais próxima
        # Prioriza a que estiver no mesmo ano da aba
        if dt.year == ano_aba and dt_invertido.year != ano_aba:
            return dt
        if dt_invertido.year == ano_aba and dt.year != ano_aba:
            return dt_invertido
        
        # Se ambas estão no mesmo ano, mantém a original
        return dt
    
    def _validar_data_retorno(self, dt_retorno: datetime, dt_saida: datetime, mes_aba: int, ano_aba: int) -> datetime:
        """
        Valida e corrige data de retorno baseado no mês da aba.
        
        Regra: Se a aba é de um mês X, o retorno pode ser no máximo no mês X+1
        (30 dias de férias). Se está em um mês posterior, provavelmente
        o dia/mês estão invertidos.
        
        Args:
            dt_retorno: Data de retorno a validar
            dt_saida: Data de saída (referência)
            mes_aba: Mês da aba (1-12)
            ano_aba: Ano da aba
            
        Returns:
            Data de retorno corrigida ou original
        """
        if not isinstance(dt_retorno, datetime) or not mes_aba:
            return dt_retorno
        
        dia, mes = dt_retorno.day, dt_retorno.month
        
        # Se dia > 12, não tem como inverter
        if dia > 12:
            return dt_retorno
        
        # Se dia == mes, não faz diferença inverter
        if dia == mes:
            return dt_retorno
        
        # Calcula o mês máximo permitido para retorno (mês da aba + 1)
        mes_max_retorno = mes_aba + 1 if mes_aba < 12 else 1
        ano_max_retorno = ano_aba if mes_aba < 12 else ano_aba + 1
        
        # Se o mês do retorno é maior que o permitido, tenta inverter
        # Ex: aba JANEIRO (1), retorno máximo FEVEREIRO (2)
        # Se retorno mostra MARÇO (3), está errado
        
        retorno_mes_ok = False
        if dt_retorno.year == ano_aba and mes <= mes_max_retorno:
            retorno_mes_ok = True
        elif dt_retorno.year == ano_max_retorno and mes <= mes_max_retorno:
            retorno_mes_ok = True
        elif mes == mes_aba:  # Mesmo mês da aba é sempre OK
            retorno_mes_ok = True
        
        if retorno_mes_ok:
            return dt_retorno
        
        # Mês do retorno parece errado, tenta inverter
        try:
            dt_invertido = datetime(dt_retorno.year, dia, mes)
        except ValueError:
            return dt_retorno
        
        # Verifica se invertido faz sentido (após saída e no mês correto)
        if dt_saida and dt_invertido < dt_saida:
            return dt_retorno  # Invertido é antes da saída, mantém original
        
        # Verifica se o mês invertido está no range permitido
        invertido_mes_ok = False
        if dt_invertido.year == ano_aba and dt_invertido.month <= mes_max_retorno:
            invertido_mes_ok = True
        elif dt_invertido.year == ano_max_retorno and dt_invertido.month <= mes_max_retorno:
            invertido_mes_ok = True
        elif dt_invertido.month == mes_aba:
            invertido_mes_ok = True
        
        if invertido_mes_ok:
            return dt_invertido
        
        return dt_retorno
    
    def _parse_data(self, valor: Any) -> Optional[datetime]:
        """Converte valor para datetime."""
        if pd.isna(valor) or str(valor).strip() in ["", "-", "nan"]:
            return None
        
        if isinstance(valor, datetime):
            return valor
        
        valor_str = str(valor).strip()
        formatos = ['%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y']
        
        for fmt in formatos:
            try:
                return datetime.strptime(valor_str.split()[0], fmt.split()[0])
            except:
                continue
        
        return None
    
    def _mapear_status(self, valor: Any) -> str:
        """Mapeia valor para status padronizado."""
        if pd.isna(valor) or str(valor).strip() in ["", "nan"]:
            return "PENDENTE"
        
        valor_str = str(valor).upper().strip()
        
        if valor_str in ["BLOQUEADO", "BLOQ"]:
            return "BLOQUEADO"
        if valor_str in ["LIBERADO", "LIB"]:
            return "LIBERADO"
        
        # Carrega padrões de "sem acesso" da configuração
        padroes_sem_acesso_str = settings.PADROES_SEM_ACESSO if hasattr(settings, 'PADROES_SEM_ACESSO') else "N/P,N\\A,NA,N/A,NP,-,NB"
        padroes_sem_acesso = [p.strip().upper() for p in padroes_sem_acesso_str.split(",")]
        
        if valor_str in padroes_sem_acesso:
            return "NA"
        
        return "PENDENTE"
    
    def _extrair_mes_ano(self, nome_aba: str) -> Tuple[Optional[int], Optional[int]]:
        """Extrai mês e ano do nome da aba."""
        meses = {
            "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3,
            "ABRIL": 4, "MAIO": 5, "JUNHO": 6, "JULHO": 7,
            "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10,
            "NOVEMBRO": 11, "DEZEMBRO": 12
        }
        
        nome_upper = nome_aba.upper()
        
        for mes_nome, mes_num in meses.items():
            if mes_nome in nome_upper:
                # Procura ano
                ano_match = re.search(r'(20\d{2})', nome_aba)
                if ano_match:
                    return (mes_num, int(ano_match.group(1)))
                
                # Formato .XX
                ano_match = re.search(r'\.(\d{2})$', nome_aba)
                if ano_match:
                    return (mes_num, 2000 + int(ano_match.group(1)))
                
                return (mes_num, datetime.now().year)
        
        return (None, None)
    
    def processar_planilha(self) -> List[Dict]:
        """Processa a planilha e extrai dados."""
        if not self.arquivo_excel or not self.arquivo_excel.exists():
            print("❌ Nenhum arquivo para processar")
            return []
        
        print(f"\n📊 Processando: {self.arquivo_excel.name}")
        
        try:
            wb = openpyxl.load_workbook(self.arquivo_excel, data_only=True)
        except Exception as e:
            print(f"   ❌ Erro: {e}")
            return []
        
        self.dados_processados = []
        self.abas_processadas = []
        
        print(f"\n📋 Total de abas na planilha: {len(wb.sheetnames)}")
        
        for nome_aba in wb.sheetnames:
            mes, ano = self._extrair_mes_ano(nome_aba)
            
            # Se não conseguir extrair mês/ano, usa valores padrão
            if mes is None:
                print(f"   ⚠️  {nome_aba}: sem mês/ano identificável, usando mês/ano atuais")
                mes = datetime.now().month
                ano = datetime.now().year
            
            ws = wb[nome_aba]
            funcionarios = self._processar_aba(ws, nome_aba, mes, ano)
            
            if funcionarios:
                self.dados_processados.extend(funcionarios)
                self.abas_processadas.append({
                    "nome": nome_aba,
                    "mes": mes,
                    "ano": ano,
                    "total_funcionarios": len(funcionarios)
                })
                print(f"   ✅ {nome_aba}: {len(funcionarios)} funcionários")
            else:
                print(f"   ⚠️  {nome_aba}: 0 funcionários (aba vazia ou sem dados válidos)")
                # Adiciona aba mesmo vazia para contar
                self.abas_processadas.append({
                    "nome": nome_aba,
                    "mes": mes,
                    "ano": ano,
                    "total_funcionarios": 0
                })
        
        print(f"\n📈 Total: {len(self.dados_processados)} funcionários, {len(self.abas_processadas)} abas")
        return self.dados_processados
    
    def _processar_aba(self, ws, nome_aba: str, mes: int, ano: int) -> List[Dict]:
        """Processa uma aba específica."""
        funcionarios = []
        
        # Mapeia colunas pelo nome do header
        colunas = {}
        colunas_por_nome = {}  # Mapeamento nome -> índice
        for idx, cell in enumerate(ws[1], start=0):
            if cell.value:
                nome_col = str(cell.value).strip()
                colunas[idx] = nome_col
                colunas_por_nome[nome_col.upper()] = idx
        
        # Encontra índices das colunas principais DINAMICAMENTE
        idx_unidade = None
        idx_nome = None
        idx_motivo = None
        idx_saida = None
        idx_retorno = None
        idx_gestor = None
        
        for nome_col, idx in colunas_por_nome.items():
            if nome_col in ["RESP.", "RESP", "UNIDADE", "RESPONSAVEL", "RESPONSÁVEL"]:
                idx_unidade = idx
            elif nome_col in ["NOME", "FUNCIONÁRIO", "FUNCIONARIO", "COLABORADOR"]:
                idx_nome = idx
            elif nome_col in ["MOTIVO", "TIPO", "RAZÃO", "RAZAO"]:
                idx_motivo = idx
            elif nome_col in ["SAÍDA", "SAIDA", "DATA SAÍDA", "DATA SAIDA", "INÍCIO", "INICIO"]:
                idx_saida = idx
            elif nome_col in ["RETORNO", "RETORNO/LIBERAÇÃO", "RETORNO/LIBERACAO", "LIBERAÇÃO", "LIBERACAO", "DATA RETORNO", "FIM"]:
                idx_retorno = idx
            elif "GESTOR" in nome_col:
                idx_gestor = idx
        
        # Fallback para índices fixos se não encontrar pelo nome
        if idx_unidade is None:
            idx_unidade = 0
        if idx_nome is None:
            idx_nome = 1
        if idx_motivo is None:
            idx_motivo = 2
        if idx_saida is None:
            idx_saida = 3
        if idx_retorno is None:
            # Procura a primeira coluna após saída que tenha dados de data
            idx_retorno = 4
            for idx_col in range(idx_saida + 1, len(colunas) + 2):
                if idx_col in colunas_por_nome.values():
                    # Verifica se o nome parece ser retorno
                    for nome, idx in colunas_por_nome.items():
                        if idx == idx_col and any(x in nome for x in ["RETORNO", "LIBERAÇÃO", "LIBERACAO", "FIM"]):
                            idx_retorno = idx_col
                            break
                    break
        
        # Índices de sistemas
        idx_sistemas = {}
        for idx, nome_col in colunas.items():
            nome_upper = nome_col.upper()
            for sistema in settings.SISTEMAS_ACESSO:
                if sistema.upper() in nome_upper:
                    idx_sistemas[sistema] = idx
                    break
        
        # Processa linhas
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                # Extração de dados brutos usando índices dinâmicos
                unidade = row[idx_unidade].value if len(row) > idx_unidade else None
                nome_bruto = row[idx_nome].value if len(row) > idx_nome else None
                motivo = row[idx_motivo].value if len(row) > idx_motivo else None
                saida_raw = row[idx_saida].value if len(row) > idx_saida else None
                retorno_raw = row[idx_retorno].value if len(row) > idx_retorno else None

                # Lógica de pular linha
                if not nome_bruto or str(nome_bruto).strip().lower() in ["", "nan", "none"]:
                    continue
                
                nome = str(nome_bruto).strip()

                # Datas
                data_saida = self._parse_data(saida_raw)
                data_retorno = self._parse_data(retorno_raw)
                
                # Valida e corrige data de SAÍDA baseado no contexto da aba
                # Ex: "12/01/2025" numa aba de DEZEMBRO pode ser na verdade "01/12/2025"
                if data_saida:
                    data_saida = self._validar_data_contexto(data_saida, mes, ano)
                
                # Correção de data de retorno se necessário (formato datetime do Excel)
                if isinstance(retorno_raw, datetime) and isinstance(data_saida, datetime):
                    data_retorno = self._corrigir_data(retorno_raw, data_saida)
                
                # Valida e corrige data de RETORNO baseado no mês da aba
                # Ex: aba JANEIRO, retorno máximo FEVEREIRO. Se mostra MARÇO, inverte para 03/02
                if data_retorno:
                    data_retorno = self._validar_data_retorno(data_retorno, data_saida, mes, ano)

                if not data_saida or not data_retorno:
                    continue

                # Gestor
                gestor = ""
                if idx_gestor and len(row) > idx_gestor:
                    gestor = str(row[idx_gestor].value or "").strip()
                    if gestor.lower() == "nan":
                        gestor = ""
                
                # Acessos
                acessos = {}
                for sistema in settings.SISTEMAS_ACESSO:
                    if sistema in idx_sistemas:
                        idx = idx_sistemas[sistema]
                        if len(row) > idx:
                            acessos[sistema] = self._mapear_status(row[idx].value)
                        else:
                            acessos[sistema] = "PENDENTE"
                    else:
                        acessos[sistema] = "NA"
                
                # Monta registro
                funcionarios.append({
                    "nome": nome,
                    "unidade": str(unidade or "").strip() if str(unidade).lower() != "nan" else "",
                    "motivo": str(motivo or "").strip() if str(motivo).lower() != "nan" else "",
                    "data_saida": data_saida.strftime('%Y-%m-%d'),
                    "data_retorno": data_retorno.strftime('%Y-%m-%d'),
                    "gestor": gestor,
                    "aba_origem": nome_aba,
                    "mes": mes,
                    "ano": ano,
                    "acessos": acessos
                })
                
            except Exception:
                continue
        
        return funcionarios

    # ==================== SINCRONIZAÇÃO ====================
    
    def sincronizar(self, forcar: bool = False) -> Dict:
        """
        Executa sincronização completa.
        
        Args:
            forcar: Se True, força download e processamento
            
        Returns:
            Dicionário com resultado da sincronização
        """
        print("=" * 60)
        print("🔄 SINCRONIZAÇÃO DE DADOS")
        print(f"⏰ {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}")
        print("=" * 60)
        
        # 1. Baixar planilha
        if not self.baixar_planilha(forcar=forcar):
            return {
                "status": "error",
                "message": "Falha no download",
                "registros": 0
            }
        
        # 2. Verificar hash
        novo_hash = self.calcular_hash()
        if not forcar and not self.arquivo_mudou(novo_hash):
            print("\n⏭️  Arquivo não foi alterado. Pulando processamento.")
            return {
                "status": "skipped",
                "message": "Arquivo não foi alterado",
                "registros": 0
            }
        
        # 3. Processar dados
        if not self.processar_planilha():
            return {
                "status": "error",
                "message": "Falha no processamento",
                "registros": 0
            }
        
        # 4. Salvar no banco
        print("\n💾 Salvando no banco de dados...")
        self.db.limpar_dados()
        total = self.db.salvar_funcionarios(self.dados_processados)
        self.db.salvar_abas(self.abas_processadas)
        
        # 5. Registrar sync
        self.db.registrar_sync(
            total_registros=total,
            total_abas=len(self.abas_processadas),
            status="SUCCESS",
            mensagem="Sincronização concluída com sucesso",
            arquivo_hash=novo_hash
        )
        
        # 6. Salvar hash
        self.salvar_hash(novo_hash)
        
        print("\n" + "=" * 60)
        print("✅ SINCRONIZAÇÃO CONCLUÍDA!")
        print(f"   📊 {total} funcionários salvos")
        print(f"   📑 {len(self.abas_processadas)} abas processadas")
        print("=" * 60)
        
        return {
            "status": "success",
            "message": f"Sincronizados {total} funcionários",
            "registros": total,
            "abas": len(self.abas_processadas),
            "timestamp": datetime.now().isoformat()
        }


# ==================== CLI ====================

def main():
    """Função principal para execução via CLI."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Sincronizador de Férias')
    parser.add_argument('--forcar', '-f', action='store_true',
                       help='Força download mesmo com cache válido')
    
    args = parser.parse_args()
    
    sync = SyncManager()
    resultado = sync.sincronizar(forcar=args.forcar)
    
    return 0 if resultado["status"] == "success" else 1


if __name__ == "__main__":
    exit(main())






